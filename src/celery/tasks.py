import os

import openpyxl
from dotenv import load_dotenv

from celery import Celery

load_dotenv()

RABBIT_HOST = os.getenv("RABBIT_HOST", "localhost")
RABBIT_USER = os.getenv("RABBIT_USER", "guest")
RABBIT_PASSWORD = os.getenv("RABBIT_PASSWORD", "guest")
RABBIT_PORT = os.getenv("RABBIT_PORT", 5672)

RABBITMQ_URL = f"amqp://{RABBIT_USER}:{RABBIT_PASSWORD}@{RABBIT_HOST}:{RABBIT_PORT}"

app_celery = Celery("tasks", broker=RABBITMQ_URL, backend="rpc://")


@app_celery.task(name="create_excel", track_started=True)
def create_excel(data):
    id = app_celery.current_task.request.id
    wb = openpyxl.Workbook()
    list = wb.active
    list.column_dimensions["A"].width = 10
    list.column_dimensions["B"].width = 20
    list.column_dimensions["C"].width = 20
    list.column_dimensions["D"].width = 20
    list.column_dimensions["E"].width = 50
    list.column_dimensions["F"].width = 15
    row = 1
    for menu in data["menus"]:
        list.cell(row=row, column=1, value=menu["id"])
        list.cell(row=row, column=2, value=menu["title"])
        list.cell(row=row, column=3, value=menu["description"])
        for submenu in data["submenus"]:
            if submenu["menu_id"] == menu["id"]:
                row += 1
                list.cell(row=row, column=2, value=submenu["id"])
                list.cell(row=row, column=3, value=submenu["title"])
                list.cell(row=row, column=4, value=submenu["description"])
                for dish in data["dishes"]:
                    if dish["submenu_id"] == submenu["id"]:
                        row += 1
                        list.cell(row=row, column=3, value=dish["id"])
                        list.cell(row=row, column=4, value=dish["title"])
                        list.cell(row=row, column=5, value=dish["description"])
                        list.cell(row=row, column=6, value=dish["price"])
        row += 1

    wb.save(f"/uploads/{id}.xlsx")
    wb.close()
